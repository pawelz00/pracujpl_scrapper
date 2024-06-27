import json
import os
import time
import requests
import re
import pandas as pd
import random
import pprint
from openpyxl import load_workbook
from datetime import datetime
from openpyxl import Workbook
from pandas import ExcelWriter, DataFrame

EXCEL_COLUMNS = {'offers': 'ID', 'lastPublicated': 'Opublikowane', 'companyName': 'Firma', 'jobTitle': 'Stanowisko',
                  'salaryDisplayText': 'Stawka', 'positionLevels': 'Poziom',
                 'url': 'Link'}

PROPERTIES_TO_KEEP_IN_EXCEL_FILE = ['jobTitle', 'companyName',
                                    'lastPublicated', 'salaryDisplayText', 'positionLevels', 'offers']
SHEET_NAME = 'Oferty pracy'
SKIP_COMPANIES = ['MindPal']
def scrap_data():
    pass
    url = 'https://it.pracuj.pl/praca?et=1%2C17%2C4&pn=1&itth=33'
    agent = 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36'
    response = requests.get(url, headers={'User-Agent': agent})
    data = response.text
    return data


def parse_data(data: str):
    pass
    regex_pattern = r'<script\s+id="__NEXT_DATA__"\s+type="application/json">(.*?)</script>'
    match = re.search(regex_pattern, data)
    not_filtered_data = json.loads(match.group(1))[
        'props']['pageProps']['data']['jobOffers']['groupedOffers']

    filtered_data = [{key: offer[key] for key in PROPERTIES_TO_KEEP_IN_EXCEL_FILE}
                     for offer in not_filtered_data]

    # Filtering out companies
    filtered_data = [offer for offer in filtered_data if offer['companyName'] not in SKIP_COMPANIES]

    for item in filtered_data:
        item['lastPublicated'] = str(item['lastPublicated']).split('T')[0]
        item['positionLevels'] = str(item['positionLevels']).replace(
            '[', '').replace(']', '').replace("'", '')
        item['url'] = item['offers'][0]['offerAbsoluteUri']
        item['offers'] = item['offers'][0]['partitionId']

    return filtered_data


def adjust_column_width(workbook: Workbook, file_path: str):
    worksheet = workbook.active

    for column in worksheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                cell_length = len(str(cell.value))
                if cell_length > max_length:
                    max_length = cell_length
            except:
                pass
                print(f"Error: {cell.value}")
        worksheet.column_dimensions[column_letter].width = max_length + 1
    workbook.save(file_path)
    workbook.close()


def format_or_create_excel_table(writer: ExcelWriter, df: pd.DataFrame):
    workbook = writer.book
    worksheet = writer.sheets['Sheet1']
    (max_row, max_col) = df.shape
    column_settings = []
    for header in df.columns:
        column_settings.append({'header': header})
    worksheet.add_table(0, 0, max_row, max_col - 1, {'columns': column_settings})
    writer.close()



def insert_rows_with_data(file_path, sheet_name, start_row, num_rows, data):
    # Load the workbook
    wb = load_workbook(filename=file_path)

    # Select the worksheet
    ws = wb[sheet_name]

    # Shift existing rows downwards
    ws.insert_rows(start_row, amount=num_rows)

    # Write data into the newly inserted rows
    for row_index, row_data in enumerate(data, start=start_row):
        for col_index, cell_value in enumerate(row_data, start=1):
            ws.cell(row=row_index, column=col_index, value=cell_value)

    # Save the workbook
    wb.save(file_path)


def save_new_data(file_path: str, data: list):
    pass

    def object_with_id_exists(objects, target_id: str, id_key: str):
        if len(objects) == 0:
            return False
        del objects[0]
        return any(obj for obj in objects if int(obj[id_key]) == int(target_id))

    def format_list(json_list, property_order):
        reordered_list = []
        for obj in json_list:
            reordered_obj = {key: obj[key] for key in property_order if key in obj}
            reordered_list.append([reordered_obj[key] for key in property_order])
        return reordered_list

    workbook = load_workbook(file_path)
    worksheet = workbook.active
    data_to_pass = data.copy()
    ids_to_retain = set()

    for column in worksheet:
        id = column[0].value
        if id is None or id == 'ID':
            continue
        object_exists = object_with_id_exists(data, id, 'offers')
        if object_exists:
            ids_to_retain.add(int(id))
            continue


    data_to_pass = [obj for obj in data_to_pass if int(obj['offers']) not in ids_to_retain]

    if len(data_to_pass) == 0:
        print("No new data to save!")
        return

    start_row = 2
    number_of_rows = len(data_to_pass)
    data_to_insert = format_list(data_to_pass, list(EXCEL_COLUMNS.keys()))
    insert_rows_with_data(file_path, SHEET_NAME, start_row, number_of_rows, data_to_insert)

    print(f"Succesfully saved {len(data_to_pass)} new rows!")


def new_excel_file_save(parsed_data: list, file_path: str):
    pass
    df = pd.DataFrame(parsed_data)
    df.rename(columns=EXCEL_COLUMNS, inplace=True)
    writer = pd.ExcelWriter(file_path, engine='xlsxwriter')
    df.to_excel(writer, sheet_name="Sheet1", startrow=1, header=False, index=False)
    format_or_create_excel_table(writer, df)



def main():
    # Getting data
    data_from_site = scrap_data()

    # Parsing data
    parsed_data = parse_data(data_from_site)

    # Checking
    folder_path = os.path.join(os.path.expanduser('~'), 'Desktop')
    excel_file_name = 'OfertyPracy.xlsx'
    file_path = os.path.join(folder_path, excel_file_name)
    file_exists = os.path.exists(file_path)

    # Saving data to excel file
    try:
        if file_exists:
            pass
            save_new_data(file_path, parsed_data)
            print("New data saved successfully!")
        else:
            pass
            new_excel_file_save(parsed_data, file_path)
            print("Data saved successfully!")
    except Exception as e:
        print(f"Error: {e}")
        return

    # Adjusting column width
    workbook = load_workbook(file_path)
    adjust_column_width(workbook, file_path)

    return 0


if __name__ == '__main__':
    main()
