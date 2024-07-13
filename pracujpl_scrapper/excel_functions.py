import pandas as pd
from openpyxl.reader.excel import load_workbook

from config import EXCEL_COLUMNS, SHEET_NAME
from pracujpl_scrapper.helpers import format_list


def adjust_column_width(file_path: str) -> None:
    workbook = load_workbook(file_path)
    worksheet = workbook.active

    for column in worksheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                cell_length = len(str(cell.value))
                if cell_length > max_length:
                    max_length = cell_length
            except:
                pass
                print(f"Error: {cell.value}")
        worksheet.column_dimensions[column_letter].width = max_length + 1
    workbook.save(file_path)
    workbook.close()


def format_or_create_excel_table(worksheet, df: pd.DataFrame) -> None:
    (max_row, max_col) = df.shape
    column_settings = []
    for header in df.columns:
        column_settings.append({'header': header})
    worksheet.add_table(0, 0, max_row, max_col - 1, {'columns': column_settings})


def insert_rows_with_data(file_path: str, sheet_name: str, start_row: int, num_rows: int, data: list) -> None:
    wb = load_workbook(filename=file_path)
    ws = wb[sheet_name]
    ws.insert_rows(start_row, amount=num_rows)

    for row_index, row_data in enumerate(data, start=start_row):
        for col_index, cell_value in enumerate(row_data, start=1):
            ws.cell(row=row_index, column=col_index, value=cell_value)

    wb.save(file_path)
    wb.close()


def save_new_data(file_path: str, data: list) -> None or bool:
    pass

    workbook = load_workbook(file_path)
    worksheet = workbook.active
    ids_in_excel = set([int(column[0].value) for column in worksheet if
                        column[0].value is not None and column[0].value != 'ID'])
    ids_in_data = set([int(obj['offers']) for obj in data])
    ids_to_retain = ids_in_data.difference(ids_in_excel)

    data_to_pass = [obj for obj in data if int(obj['offers']) in ids_to_retain]

    if len(data_to_pass) == 0:
        print("No new data to save!")
        return False

    start_row = 2
    number_of_rows = len(data_to_pass)
    data_to_insert = format_list(data_to_pass, list(EXCEL_COLUMNS.keys()))
    insert_rows_with_data(file_path, SHEET_NAME, start_row, number_of_rows, data_to_insert)

    new_df = pd.read_excel(file_path)
    new_df.rename(columns=EXCEL_COLUMNS, inplace=True)

    with pd.ExcelWriter(file_path, engine='xlsxwriter') as writer:
        new_df.to_excel(writer, sheet_name=SHEET_NAME, index=False)
        worksheet = writer.sheets[SHEET_NAME]
        format_or_create_excel_table(worksheet, new_df)

    print(f"Successfully saved {len(data_to_pass)} new rows!")
    return True


def new_excel_file_save(parsed_data: list, file_path: str) -> None:
    pass
    
    df = pd.DataFrame(parsed_data)
    df.rename(columns=EXCEL_COLUMNS, inplace=True)

    with pd.ExcelWriter(file_path, engine='xlsxwriter') as writer:
        df.to_excel(writer, sheet_name=SHEET_NAME, index=False)
        worksheet = writer.sheets[SHEET_NAME]
        format_or_create_excel_table(worksheet, df)
